from io import BytesIO

from openpyxl import load_workbook

from app_event.models import Event


def test_export_xlsx_superuser(client, superuser_headers, published_event):
    response = client.get("/v1/event/export/xlsx", headers=superuser_headers)
    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="events.xlsx"' in response["Content-Disposition"]

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) >= 2
    headers = rows[0]
    assert "Название" in headers
    assert "Описание" in headers

    data_row = rows[1]
    assert published_event.name in data_row


def test_export_xlsx_regular_user_forbidden(client, regular_user_headers):
    response = client.get("/v1/event/export/xlsx", headers=regular_user_headers)
    assert response.status_code == 403


def test_export_xlsx_unauthenticated(client):
    response = client.get("/v1/event/export/xlsx")
    assert response.status_code == 401


def test_export_xlsx_with_filters(
    client, superuser_headers, published_event, unpublished_event, event_place, db
):
    high_rate_event = Event.objects.create(
        published=True,
        name="High Rate Export Event",
        description="desc",
        start_date="2025-06-01T10:00:00Z",
        end_date="2025-06-01T18:00:00Z",
        author="Author",
        place=event_place,
        rate=20,
        status=Event.SOON,
    )

    response = client.get(
        "/v1/event/export/xlsx?rate_gte=15", headers=superuser_headers
    )
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    names = [row[0] for row in rows[1:]]
    assert high_rate_event.name in names
    assert published_event.name not in names


def test_export_xlsx_place_filter(
    client, superuser_headers, published_event, event_place
):
    response = client.get(
        f"/v1/event/export/xlsx?place_id={event_place.id}", headers=superuser_headers
    )
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    names = [row[0] for row in rows[1:]]
    assert published_event.name in names
