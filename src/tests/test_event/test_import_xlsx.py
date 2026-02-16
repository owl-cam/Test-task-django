from datetime import datetime
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from app_event.models import Event
from app_event_place.models import EventPlace


def _create_xlsx_bytes(rows, headers=None):
    """Helper to create xlsx file bytes."""
    wb = Workbook()
    ws = wb.active

    if headers is None:
        headers = [
            "Название",
            "Описание",
            "Дата и время публикации",
            "Дата и время начала проведения",
            "Дата и время завершения проведения",
            "Название места проведения",
            "Широта",
            "Долгота",
            "Рейтинг (от 0 до 25)",
        ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_num, row_data in enumerate(rows, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _create_file(rows):
    """Create SimpleUploadedFile from rows."""
    xlsx_bytes = _create_xlsx_bytes(rows)
    return SimpleUploadedFile(
        "events.xlsx",
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def test_import_xlsx_superuser_success(client, superuser_headers, event_place, db):
    rows = [
        (
            "Imported Event",
            "Test description",
            None,
            datetime(2025, 6, 1, 10, 0),
            datetime(2025, 6, 1, 18, 0),
            event_place.name,
            None,
            None,
            10,
        ),
    ]
    file = _create_file(rows)

    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True
    assert data["imported_count"] == 1
    assert data["total_rows"] == 1
    assert len(data["errors"]) == 0

    event = Event.objects.get(name="Imported Event")
    assert event.description == "Test description"
    assert event.place == event_place


def test_import_xlsx_regular_user_forbidden(client, regular_user_headers, db):
    file = _create_file([])
    response = client.post(
        "/v1/event/import",
        headers=regular_user_headers,
        FILES={"file": file},
    )
    assert response.status_code == 403


def test_import_xlsx_unauthenticated(client, db):
    file = _create_file([])
    response = client.post(
        "/v1/event/import",
        FILES={"file": file},
    )
    assert response.status_code == 401


def test_import_xlsx_empty_file(client, superuser_headers, db):
    file = _create_file([])
    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True
    assert data["imported_count"] == 0
    assert data["total_rows"] == 0


def test_import_xlsx_missing_required_fields(client, superuser_headers, db):
    rows = [
        (
            None,
            "Description only",
            None,
            datetime(2025, 6, 1, 10, 0),
            datetime(2025, 6, 1, 18, 0),
            None,
            None,
            None,
            10,
        ),
    ]
    file = _create_file(rows)

    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is False
    assert len(data["errors"]) > 0
    assert any(e["field"] == "Название" for e in data["errors"])


def test_import_xlsx_invalid_rate(client, superuser_headers, event_place, db):
    rows = [
        (
            "Test Event",
            "Description",
            None,
            datetime(2025, 6, 1, 10, 0),
            datetime(2025, 6, 1, 18, 0),
            event_place.name,
            None,
            None,
            30,
        ),
    ]
    file = _create_file(rows)

    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is False
    assert any(e["field"] == "Рейтинг (от 0 до 25)" for e in data["errors"])


def test_import_xlsx_creates_new_place(client, superuser_headers, db):
    rows = [
        (
            "Event With New Place",
            "Description",
            None,
            datetime(2025, 6, 1, 10, 0),
            datetime(2025, 6, 1, 18, 0),
            "New Venue",
            55.7558,
            37.6173,
            10,
        ),
    ]
    file = _create_file(rows)

    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True
    assert "New Venue" in data["created_places"]

    place = EventPlace.objects.get(name="New Venue")
    assert float(place.lat) == 55.7558
    assert float(place.long) == 37.6173


def test_import_xlsx_place_without_coords(client, superuser_headers, db):
    rows = [
        (
            "Event",
            "Description",
            None,
            datetime(2025, 6, 1, 10, 0),
            datetime(2025, 6, 1, 18, 0),
            "Unknown Place",
            None,
            None,
            10,
        ),
    ]
    file = _create_file(rows)

    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is False
    assert any(e["field"] == "Название места проведения" for e in data["errors"])


def test_import_xlsx_invalid_date_format(client, superuser_headers, db):
    rows = [
        (
            "Event",
            "Description",
            None,
            "not-a-date",
            datetime(2025, 6, 1, 18, 0),
            None,
            None,
            None,
            10,
        ),
    ]
    file = _create_file(rows)

    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is False
    assert any(e["field"] == "Дата и время начала проведения" for e in data["errors"])


def test_import_xlsx_multiple_events(client, superuser_headers, event_place, db):
    rows = [
        (
            "Event 1",
            "Description 1",
            None,
            datetime(2025, 6, 1, 10, 0),
            datetime(2025, 6, 1, 18, 0),
            event_place.name,
            None,
            None,
            5,
        ),
        (
            "Event 2",
            "Description 2",
            None,
            datetime(2025, 7, 1, 10, 0),
            datetime(2025, 7, 1, 18, 0),
            event_place.name,
            None,
            None,
            15,
        ),
    ]
    file = _create_file(rows)

    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True
    assert data["imported_count"] == 2
    assert data["total_rows"] == 2

    assert Event.objects.filter(name="Event 1").exists()
    assert Event.objects.filter(name="Event 2").exists()


def test_import_xlsx_invalid_file(client, superuser_headers, db):
    file = SimpleUploadedFile(
        "events.xlsx",
        b"not an xlsx file",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response = client.post(
        "/v1/event/import",
        headers=superuser_headers,
        FILES={"file": file},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["success"] is False
    assert len(data["errors"]) > 0
    assert any(e["field"] == "file" for e in data["errors"])
