from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from app_event.models import Event
from app_event_place.models import EventPlace


class EventXlsxService:
    IMPORT_COLUMNS = [
        "Название",
        "Описание",
        "Дата и время публикации",
        "Дата и время начала проведения",
        "Дата и время завершения проведения",
        "Название места проведения",
        "Широта",
        "Долгота",
        "Рейтинг (от 0 до 25)",
    ]

    EXPORT_COLUMNS = IMPORT_COLUMNS + [
        "Автор",
        "Статус",
        "Опубликовано",
    ]

    STATUS_DISPLAY = {
        Event.SOON: "Скоро",
        Event.ONGOING: "Идет",
        Event.OVER: "Закончилось",
    }

    STATUS_REVERSE = {v: k for k, v in STATUS_DISPLAY.items()}

    @staticmethod
    def _make_naive(dt: datetime | None) -> datetime | None:
        if dt is None:
            return None
        return dt.replace(tzinfo=None)

    def export_events(self, filters: dict) -> BytesIO:
        queryset = Event.objects.select_related("place").all()

        if filters.get("publish_date_gte"):
            queryset = queryset.filter(publish_date__gte=filters["publish_date_gte"])
        if filters.get("publish_date_lte"):
            queryset = queryset.filter(publish_date__lte=filters["publish_date_lte"])
        if filters.get("start_date_gte"):
            queryset = queryset.filter(start_date__gte=filters["start_date_gte"])
        if filters.get("start_date_lte"):
            queryset = queryset.filter(start_date__lte=filters["start_date_lte"])
        if filters.get("end_date_gte"):
            queryset = queryset.filter(end_date__gte=filters["end_date_gte"])
        if filters.get("end_date_lte"):
            queryset = queryset.filter(end_date__lte=filters["end_date_lte"])
        if filters.get("place_id"):
            queryset = queryset.filter(place_id=filters["place_id"])
        if filters.get("rate_gte") is not None:
            queryset = queryset.filter(rate__gte=filters["rate_gte"])
        if filters.get("rate_lte") is not None:
            queryset = queryset.filter(rate__lte=filters["rate_lte"])

        wb = Workbook()
        ws = wb.active
        ws.title = "События"

        for col_num, header in enumerate(self.EXPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = cell.font.copy(bold=True)

        for row_num, event in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=event.name)
            ws.cell(row=row_num, column=2, value=event.description)
            ws.cell(row=row_num, column=3, value=self._make_naive(event.publish_date))
            ws.cell(row=row_num, column=4, value=self._make_naive(event.start_date))
            ws.cell(row=row_num, column=5, value=self._make_naive(event.end_date))
            ws.cell(
                row=row_num, column=6, value=event.place.name if event.place else None
            )
            ws.cell(
                row=row_num,
                column=7,
                value=float(event.place.lat) if event.place else None,
            )
            ws.cell(
                row=row_num,
                column=8,
                value=float(event.place.long) if event.place else None,
            )
            ws.cell(row=row_num, column=9, value=event.rate)
            ws.cell(row=row_num, column=10, value=event.author)
            ws.cell(
                row=row_num,
                column=11,
                value=self.STATUS_DISPLAY.get(event.status, event.status),
            )
            ws.cell(row=row_num, column=12, value="Да" if event.published else "Нет")

        for col_num in range(1, len(self.EXPORT_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def import_events(self, file_data: bytes, author: str) -> dict:
        errors = []
        created_places = []
        imported_count = 0

        try:
            wb = load_workbook(filename=BytesIO(file_data))
        except Exception as e:
            return {
                "success": False,
                "total_rows": 0,
                "imported_count": 0,
                "errors": [
                    {
                        "row": 0,
                        "field": "file",
                        "error": f"Ошибка чтения файла: {e}",
                        "value": None,
                    }
                ],
                "created_places": [],
            }

        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Filter out completely empty rows
        rows = [(i, row) for i, row in enumerate(all_rows, 2) if any(cell is not None for cell in row)]
        total_rows = len(rows)

        if total_rows == 0:
            return {
                "success": True,
                "total_rows": 0,
                "imported_count": 0,
                "errors": [],
                "created_places": [],
            }

        events_to_create = []
        places_cache = {p.name: p for p in EventPlace.objects.all()}

        for row_num, row in rows:
            row_errors = []

            name = row[0] if len(row) > 0 else None
            description = row[1] if len(row) > 1 else None
            publish_date = row[2] if len(row) > 2 else None
            start_date = row[3] if len(row) > 3 else None
            end_date = row[4] if len(row) > 4 else None
            place_name = row[5] if len(row) > 5 else None
            lat = row[6] if len(row) > 6 else None
            long = row[7] if len(row) > 7 else None
            rate = row[8] if len(row) > 8 else None

            if not name:
                row_errors.append(
                    {
                        "row": row_num,
                        "field": "Название",
                        "error": "Обязательное поле",
                        "value": None,
                    }
                )

            if not description:
                row_errors.append(
                    {
                        "row": row_num,
                        "field": "Описание",
                        "error": "Обязательное поле",
                        "value": None,
                    }
                )

            if not start_date:
                row_errors.append(
                    {
                        "row": row_num,
                        "field": "Дата и время начала проведения",
                        "error": "Обязательное поле",
                        "value": None,
                    }
                )
            elif not isinstance(start_date, datetime):
                row_errors.append(
                    {
                        "row": row_num,
                        "field": "Дата и время начала проведения",
                        "error": "Неверный формат даты",
                        "value": str(start_date),
                    }
                )

            if not end_date:
                row_errors.append(
                    {
                        "row": row_num,
                        "field": "Дата и время завершения проведения",
                        "error": "Обязательное поле",
                        "value": None,
                    }
                )
            elif not isinstance(end_date, datetime):
                row_errors.append(
                    {
                        "row": row_num,
                        "field": "Дата и время завершения проведения",
                        "error": "Неверный формат даты",
                        "value": str(end_date),
                    }
                )

            if publish_date and not isinstance(publish_date, datetime):
                row_errors.append(
                    {
                        "row": row_num,
                        "field": "Дата и время публикации",
                        "error": "Неверный формат даты",
                        "value": str(publish_date),
                    }
                )
                publish_date = None

            if rate is None:
                row_errors.append(
                    {
                        "row": row_num,
                        "field": "Рейтинг (от 0 до 25)",
                        "error": "Обязательное поле",
                        "value": None,
                    }
                )
            else:
                try:
                    rate = int(rate)
                    if rate < 0 or rate > 25:
                        row_errors.append(
                            {
                                "row": row_num,
                                "field": "Рейтинг (от 0 до 25)",
                                "error": "Рейтинг должен быть от 0 до 25",
                                "value": str(rate),
                            }
                        )
                except (ValueError, TypeError):
                    row_errors.append(
                        {
                            "row": row_num,
                            "field": "Рейтинг (от 0 до 25)",
                            "error": "Должно быть числом",
                            "value": str(rate),
                        }
                    )

            place = None
            if place_name:
                place_name = str(place_name).strip()
                if place_name in places_cache:
                    place = places_cache[place_name]
                else:
                    if lat is not None and long is not None:
                        try:
                            lat_decimal = Decimal(str(lat))
                            long_decimal = Decimal(str(long))
                            place = EventPlace(
                                name=place_name, lat=lat_decimal, long=long_decimal
                            )
                            places_cache[place_name] = place
                            created_places.append(place_name)
                        except (InvalidOperation, ValueError, TypeError) as e:
                            row_errors.append(
                                {
                                    "row": row_num,
                                    "field": "Широта/Долгота",
                                    "error": f"Неверный формат координат: {e}",
                                    "value": f"lat={lat}, long={long}",
                                }
                            )
                    else:
                        row_errors.append(
                            {
                                "row": row_num,
                                "field": "Название места проведения",
                                "error": "Место не найдено, а координаты не указаны",
                                "value": place_name,
                            }
                        )

            if row_errors:
                errors.extend(row_errors)
                continue

            events_to_create.append(
                {
                    "name": name,
                    "description": description,
                    "publish_date": publish_date,
                    "start_date": start_date,
                    "end_date": end_date,
                    "place": place,
                    "rate": rate,
                    "author": author,
                    "published": False,
                    "status": Event.SOON,
                }
            )

        if errors:
            return {
                "success": False,
                "total_rows": total_rows,
                "imported_count": 0,
                "errors": errors,
                "created_places": [],
            }

        try:
            with transaction.atomic():
                new_places = [p for p in places_cache.values() if p.pk is None]
                EventPlace.objects.bulk_create(new_places)

                for p in new_places:
                    places_cache[p.name] = p

                events = []
                for event_data in events_to_create:
                    place = event_data.pop("place")
                    if place:
                        event_data["place"] = places_cache[place.name]
                    events.append(Event(**event_data))

                Event.objects.bulk_create(events)
                imported_count = len(events)

        except Exception as e:
            return {
                "success": False,
                "total_rows": total_rows,
                "imported_count": 0,
                "errors": [
                    {
                        "row": 0,
                        "field": "database",
                        "error": f"Ошибка сохранения: {e}",
                        "value": None,
                    }
                ],
                "created_places": [],
            }

        return {
            "success": True,
            "total_rows": total_rows,
            "imported_count": imported_count,
            "errors": [],
            "created_places": created_places,
        }
